from datetime import datetime
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http.response import JsonResponse
import openpyxl
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from inventory.filters import CategoryFilter, ProductArticleFilter, SubCategoryFilter, ProductFilter, SectionFilter
from authentication import permissions
from .utils import save_db_snapshot
from .models import Category, Product, SubCategory, Brand, ProductColor, ProductSize
from .serializers import *
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView

from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
import json
from rest_framework import generics

from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie, vary_on_headers
from django.core.cache import cache
from django.db.models import Sum
from django.conf import settings
from django.core.files.storage import default_storage
import os
import uuid

import subprocess
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt




def generate_unique_number():
    # Get current datetime in the format YYYYMMDDHHMMSS
    current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
    # Generate a UUID and take only the first 8 characters for randomness
    unique_id = str(uuid.uuid4())[:8]
    # Combine the datetime and unique_id
    unique_number = f"{current_datetime}{unique_id}"
    return unique_number


class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.all()
    serializer_class = SectionSerializer
    permission_classes = [IsAuthenticated] 
    filter_backends = [DjangoFilterBackend]
    filterset_class = SectionFilter

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.prefetch_related('section')
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated] 
    filter_backends = [DjangoFilterBackend]
    filterset_class = CategoryFilter

class SubcategoryViewSet(viewsets.ModelViewSet):
    queryset = SubCategory.objects.prefetch_related('category')
    serializer_class = SubcategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = SubCategoryFilter

class ProductArticleViewSet(viewsets.ModelViewSet):
    queryset = ProductArticle.objects.all()
    serializer_class = ProductArticleSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = ProductArticleFilter


#unused
# class ProductVariantViewSet(viewsets.ModelViewSet):
#     queryset = ProductVariant.objects.prefetch_related('size', 'color')
#     serializer_class = ProductVariantSerializer
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend]

#     # @method_decorator(cache_page(60*60*18))
#     # def dispatch(self, request, *args, **kwargs):
#     #     return super().dispatch(request, *args, **kwargs)

# class ProductViewSet(viewsets.ModelViewSet):
#     queryset = Product.objects.all()
#     serializer_class = ProductSerializer
#     permission_classes = [IsAuthenticated]
#     filter_backends = [DjangoFilterBackend]
#     filterset_class = ProductFilter

#     parser_classes = (JSONParser, MultiPartParser, FormParser)

#     def get_queryset(self):
#         products = Product.objects.all().order_by('-updated_at')
#         return products

#     def get_serializer_class(self):
#         if self.action == 'create' or self.action == 'update':
#             return ProductWriteSerializer  # Use WriteSerializer for POST requests
#         return self.serializer_class  # Use ReadSerializer for other operations

#     def create(self, request, *args, **kwargs):
#         data = request.data
#         _mutable = data._mutable

#         # set to mutable
#         data._mutable = True
#         # Ensure nested JSON fields are parsed correctly
#         for field in ['brand', 'section', 'category', 'subcategory', 'size', 'color', 'article_no']:
#             if field in data and isinstance(data[field], str):
#                 try:
#                     data[field] = json.loads(data[field])
#                 except json.JSONDecodeError:
#                     return Response({'error': f'Invalid JSON format for {field}'}, status=status.HTTP_400_BAD_REQUEST)

#         # Extract and create or retrieve related objects
#         brand = self.get_or_create_brand(data.get('brand'))
#         section = self.get_or_create_section(data.get('section'))
#         category = self.get_or_create_category(data.get('category'), section)
#         subcategory = self.get_or_create_subcategory(data.get('subcategory'), category)
#         article = self.get_or_create_article(data.get('article_no'))
#         size = self.get_or_create_product_size(data.get('size'))
#         color = self.get_or_create_product_color(data.get('color'))

#         # Assign the primary keys of related objects to data dictionary
#         data['brand'] = brand.pk
#         data['section'] = section.pk
#         data['category'] = category.pk
#         data['article_no'] = article.pk

#         if subcategory:
#             data['subcategory'] = subcategory.pk
#         if size:
#             data['size'] = size.pk
#         if color:
#             data['color'] = color.pk

#         if article:
#             data['article_no'] = article.pk

#         # Validate and save the serializer
#         # set mutable flag back
#         data._mutable = _mutable
#         serializer = self.get_serializer(data=data)
#         serializer.is_valid(raise_exception=True)
#         self.perform_create(serializer)

#         data._mutable = True

#         data['brand'] = brand
#         data['section'] = section
#         data['category'] = category

#         if subcategory:
#             data['subcategory'] = subcategory
#         if size:
#             data['size'] = size
#         if color:
#             data['color'] = color
#         if article:
#             data['article_no'] = article

#         # Validate and save the serializer
#         # set mutable flag back
#         data._mutable = _mutable

#         # Return the response
#         headers = self.get_success_headers(serializer.data)
#         return Response(ProductSerializer(data).data, status=status.HTTP_201_CREATED, headers=headers)

#     def update(self, request, *args, **kwargs):
#         data = request.data
#         _mutable = data._mutable

#         # set to mutable
#         data._mutable = True
#         # Ensure nested JSON fields are parsed correctly
#         for field in ['brand', 'section', 'category', 'subcategory', 'size', 'color', 'article_no']:
#             if field in data and isinstance(data[field], str):
#                 try:
#                     data[field] = json.loads(data[field])
#                 except json.JSONDecodeError:
#                     return Response({'error': f'Invalid JSON format for {field}'}, status=status.HTTP_400_BAD_REQUEST)

#         # Extract and create or retrieve related objects
#         brand = self.get_or_create_brand(data.get('brand'))
#         section = self.get_or_create_section(data.get('section'))
#         category = self.get_or_create_category(data.get('category'), section)
#         subcategory = self.get_or_create_subcategory(data.get('subcategory'), category)
#         article = self.get_or_create_article(data.get('article_no'))
#         size = self.get_or_create_product_size(data.get('size'))
#         color = self.get_or_create_product_color(data.get('color'))

#         # Assign the primary keys of related objects to data dictionary
#         data['brand'] = brand.pk
#         data['section'] = section.pk
#         data['category'] = category.pk
        
#         if subcategory:
#             data['subcategory'] = subcategory.pk
#         if size:
#             data['size'] = size.pk
#         if color:
#             data['color'] = color.pk
#         if article:
#             data['article_no'] = article.pk

#         # Validate and save the serializer
#         # set mutable flag back
#         data._mutable = _mutable
#         partial = kwargs.pop('partial', False)
#         instance = self.get_object()
#         serializer = self.get_serializer(instance, data=data, partial=partial)
#         serializer.is_valid(raise_exception=True)
#         self.perform_update(serializer)

#         return Response(serializer.data)


#     def get_or_create_brand(self, brand_data):
#         if not brand_data:
#             return None

#         brand_name = brand_data.get('name').title()

#         try:
#             brand = Brand.objects.get(name__iexact=brand_name) 
#         except:
#             brand, created = Brand.objects.get_or_create(name=brand_name)
#         return brand
    
#     def get_or_create_section(self, section_data):
#         if not section_data:
#             return None

#         section_name = section_data.get('name').title()
#         try:
#             section = Section.objects.get(name__iexact=section_name) 
#         except:
#             section, created = Section.objects.get_or_create(name=section_name)
#         return section

#     def get_or_create_category(self, category_data, section):
#         if not category_data:
#             return None

#         category_name = category_data.get('name')
#         try:
#             category = Category.objects.get(name__iexact=category_name) 
#         except:
#             category, created = Category.objects.get_or_create(name=category_name, section=section)
#         return category

#     def get_or_create_subcategory(self, subcategory_data, category):
#         if not subcategory_data:
#             return None

#         subcategory_name = subcategory_data.get('name')
#         try:
#             subcategory = SubCategory.objects.get(name__iexact=subcategory_name) 
#         except:
#             subcategory, created = SubCategory.objects.get_or_create(name=subcategory_name, category=category)

#         return subcategory
    
#     def get_or_create_article(self, article_data):
#         print("art",article_data)
#         if not article_data:
#             return None

#         article_name = article_data.get('article')
#         try:
#             article = ProductArticle.objects.get(article__iexact=article_name) 
#         except:
#             article, created = ProductArticle.objects.get_or_create(article=article_name)

#         return article

#     def get_or_create_product_size(self, size_data):
#         if not size_data:
#             return None

#         size_name = size_data.get('size')
#         try:
#             size = ProductSize.objects.get(size__iexact=size_name)
#         except:
#             size, created = ProductSize.objects.get_or_create(size=size_name)
#         return size

#     def get_or_create_product_color(self, color_data):
#         if not color_data:
#             return None

#         color_name = color_data.get('color').title()
#         try:
#             color = ProductColor.objects.get(color__iexact=color_name)
#         except:
#             color, created = ProductColor.objects.get_or_create(color=color_name)
#         return color

class FilterView(APIView):
    permission_classes = [permissions.IsAppUser]

    def get(self, request, *args, **kwargs):
        cached_filters = cache.get('filters')

        if cached_filters:
            # Return cached cached_filters if available
            return Response(cached_filters)
        
        sections = Section.objects.all()
        categories = Category.objects.all()
        subcategories = SubCategory.objects.all()
        brands = Brand.objects.all()
        product_colors = ProductColor.objects.all()
        product_sizes = ProductSize.objects.all()
        product_articles = ProductArticle.objects.all()
        
        data = {
            'sections': SectionSerializer(sections, many=True).data,
            'categories': CategorySerializer(categories, many=True).data,
            'subcategories': SubcategorySerializer(subcategories, many=True).data,
            'brands': BrandSerializer(brands, many=True).data,
            'product_colors': ProductColorSerializer(product_colors, many=True).data,
            'product_sizes': ProductSizeSerializer(product_sizes, many=True).data,
            'product_articles' : ProductArticleSerializer(product_articles, many=True).data
        }
        
        return Response(data)
    
class NewProductVariantViewSet(viewsets.ModelViewSet):
    queryset = ProductVariant.objects.prefetch_related('size', 'color')
    serializer_class = ProductVariantSerializer
    permission_classes = [permissions.IsAppUser]

    # @method_decorator(cache_page(60*60*18))
    # def dispatch(self, request, *args, **kwargs):
    #     return super().dispatch(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        # Attempt to get the variants list from cache
        cached_products = cache.get('product_variant_list')

        if cached_products:
            # Return cached variants if available
            return Response(cached_products)

        # Fetch variants from the database if not cached
        response = super().list(request, *args, **kwargs)

        # Cache the variants for 18 Hrs
        cache.set('product_variant_list', response.data, timeout=60*60*18)
        return response

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.prefetch_related('variants', 'section', 'category', 'subcategory', 'brand', 'article_no')
    serializer_class = NewProductSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    permission_classes = [permissions.IsAppUser]

    # # ref - https://stackoverflow.com/questions/51499175/caching-a-viewset-with-drf-typeerror-wrapped-view
    # @method_decorator(cache_page(60*60*18))
    # def dispatch(self, request, *args, **kwargs):
    #     return super().dispatch(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        # Attempt to get the product list from cache
        cached_products = cache.get('product_list')
        

        if cached_products:
            # Return cached products if available=
            return Response(cached_products)

        # Fetch products from the database if not cached
        response = super().list(request, *args, **kwargs)

        # Cache the products for 18 Hrs
        cache.set('product_list', response.data, timeout=60*60*18)
        return response

class ProductByBarcodeAPIView(generics.RetrieveAPIView):
    # queryset = Product.objects.prefetch_related('variants', 'section', 'category', 'subcategory', 'brand', 'article_no')
    # serializer_class = NewProductSerializer
    permission_classes = [permissions.IsAppUser]

    serializer_class = NewProductSerializer

    def get_queryset(self):
        barcode = self.kwargs['barcode']
        return Product.objects.filter(barcode__endswith=barcode)

    def get(self, request, barcode, *args, **kwargs):
        products = self.get_queryset()
        if products.exists():
            serializer = self.get_serializer(products, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response({"detail": "No products found."}, status=status.HTTP_404_NOT_FOUND)
        
class ProductInventoryByCategory(APIView):
    def get(self, request):
        # Group by category and subcategory, and sum the inventory for all products
        data = Product.objects.values('category__name', 'subcategory__name').annotate(
            total_inventory=Sum('variants__inventory')  # Sum inventory for all variants
        ).order_by('category__name', 'subcategory__name')  # Sort by category and subcategory

        # Structure the data for the chart
        labels = [f"{item['category__name']}/{item['subcategory__name']}" for item in data]
        inventory = [item['total_inventory'] for item in data]

        # Return the structured data
        return Response({
            'labels': labels,
            'inventory': inventory
        })
    
class ProductVariantImportView(APIView):
    '''
        Deprecated
    '''
    permission_classes = [permissions.IsShopOwner]

    def post(self, request):
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)

        file = request.FILES['file']
        failed_rows = []
        transaction_ID = generate_unique_number()
        output_file_path = f"Transaction_Summary_{transaction_ID}.xlsx"

        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Create a new workbook for failed rows
            error_workbook = openpyxl.Workbook()
            error_sheet = error_workbook.active
            error_sheet.title = "Failed Rows"

            # Copy header row from the original sheet and add a "Failure Reason" column
            headers = [cell.value for cell in sheet[1]]
            headers.append('Failure Reason')
            error_sheet.append(headers)
            success_count = 0
            failuer_count = 0

            product_added_count = 0
            variant_added_count = 0

            product_updated_count = 0
            variant_updated_count = 0


            # Iterate over rows in the Excel file
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row

                # Check if the row is empty or contains only None values
                if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                    continue  # Skip empty rows
                product_created = False
                with transaction.atomic():
                    try:
                        row = [val.strip() if isinstance(val, str) else val for val in row]
                        (barcode, product_name, brand_name, section_name, article_name, 
                         category_name, subcategory_name, color_name, size_name, selling_price, mfd_date, isMultipack, multipack_qty,
                         inventory, isVariant, *cols) = row

                        # Retrieve or create the brand, section, category, and subcategory
                        brand = Brand.objects.get_or_create(name=str(brand_name).title())[0] if brand_name else None
                        section = Section.objects.get_or_create(name=str(section_name).title())[0] if section_name else None
                        category = Category.objects.get_or_create(name=str(category_name).title(), section=section)[0] if category_name else None
                        subcategory = SubCategory.objects.get_or_create(name=str(subcategory_name).title(), category=category)[0] if subcategory_name else None
                        article = ProductArticle.objects.get_or_create(article=str(article_name).upper())[0] if article_name else None

                        # Try to fetch the product, or create if it does not exist
                        product, product_created = Product.objects.get_or_create(
                            barcode=barcode,
                            defaults= {
                                'name' : product_name,
                                'brand': brand,
                                'section': section,
                                'category': category,
                                'subcategory': subcategory,
                                'is_multi_pack': True if isMultipack == "Yes" else False,
                                'multi_pack_quantity':  multipack_qty,
                                'article_no': article,
                            }
                        )

                        if(product_created):
                            product_added_count += 1

                        # Get or create the size and color
                        size = ProductSize.objects.get_or_create(size=str(size_name).upper())[0] if size_name else None
                        color = ProductColor.objects.get_or_create(color=str(color_name).title())[0] if color_name else None

                        # Convert mfd_date from string to date if necessary
                        if isinstance(mfd_date, str):
                            mfd_date = datetime.strptime(mfd_date, '%Y-%m-%d').date()

                        # Check for existing variant and update or create as needed
                        variant, variant_created = ProductVariant.objects.update_or_create(
                            product=product,
                            size=size,
                            color=color,
                            mfd_date=mfd_date,
                            defaults={
                                'buying_price': selling_price * 0.80,
                                'selling_price': selling_price,
                                'applicable_gst': 0,
                                'inventory': inventory * product.multi_pack_quantity if product.is_multi_pack else inventory,
                            }
                        )

                        if not variant_created:
                            variant.inventory += inventory * product.multi_pack_quantity if product.is_multi_pack else inventory
                            variant.save()
                            variant_updated_count += 1
                        else:
                            variant_added_count += 1    

                        success_count += 1

                    except ValidationError as e:
                        failure_reason = f"Validation error: {e}"
                        transaction.set_rollback(True)
                        failed_rows.append((*row, failure_reason))
                        failuer_count += 1
                        if product_created:
                            product_added_count -= 1
                    except Exception as e:
                        failure_reason = f"Error: {e}"
                        transaction.set_rollback(True)
                        failed_rows.append((*row, failure_reason))
                        failuer_count += 1
                        if product_created:
                            product_added_count -= 1

            # Write failed rows to the error workbook
            for failed_row in failed_rows:
                error_sheet.append(failed_row)

            # Save the error workbook to a file in the media folder
            failed_imports_path = os.path.join(settings.MEDIA_ROOT, output_file_path)
            error_workbook.save(failed_imports_path)
            
            # Return a JSON response with failed rows
            return Response({
                'message': 'File processed with some errors.' if failed_rows else 'File processed successfully.',
                'failed_rows': failed_rows,
                'product_added_count': product_added_count,
                'product_updated_count': product_updated_count,
                'variant_added_count': variant_added_count,
                'variant_updated_count': variant_updated_count,
                'failure_count': failuer_count,
                'success_count': success_count,
                'download_url': default_storage.url('failed_imports.xlsx'),
                'transaction_id': transaction_ID
            }, status=status.HTTP_207_MULTI_STATUS if failed_rows else status.HTTP_200_OK)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        

class ProductVariantImportViewNew(APIView):
    permission_classes = [permissions.IsShopOwner]

    def post(self, request):
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)

        file = request.FILES['file']
        failed_rows = []
        success_rows = []
        transaction_ID = generate_unique_number()
        output_file_path = f"Transaction_Summary_{transaction_ID}.xlsx"

        try:

            db_snapshot_name = save_db_snapshot(transaction_ID)

            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Create a new workbook for logging rows
            log_workbook = openpyxl.Workbook()
            error_sheet = log_workbook.active
            error_sheet.title = "Failed Rows"
            success_sheet = log_workbook.create_sheet(title="Successful Rows")

            # Copy header row from the original sheet and add a "Failure Reason" column
            headers = [cell.value for cell in sheet[1]]
            headers.append('Failure Reason')
            error_sheet.append(headers)

            # Create headers for the success log sheet with additional log columns
            success_headers = headers[:-1] + ['Product Log', 'Variant Log', 'Updated Fields']
            success_sheet.append(success_headers)

            success_count = 0
            failure_count = 0

            product_added_count = 0
            variant_added_count = 0

            product_updated_count = 0
            variant_updated_count = 0

            # Iterate over rows in the Excel file
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Check if the row is empty or contains only None values
                if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                    continue  # Skip empty rows

                product_created = False
                with transaction.atomic():
                    try:
                        row = [val.strip() if isinstance(val, str) else val for val in row]
                        (barcode, product_name, brand_name, section_name, article_name, 
                         category_name, subcategory_name, color_name, size_name, selling_price, mfd_date, isMultipack, multipack_qty,
                         inventory, isVariant, *cols) = row

                        # Retrieve or create the brand, section, category, and subcategory
                        brand = Brand.objects.get_or_create(name=str(brand_name).title())[0] if brand_name else None
                        section = Section.objects.get_or_create(name=str(section_name).title())[0] if section_name else None
                        category = Category.objects.get_or_create(name=str(category_name).title(), section=section)[0] if category_name else None
                        subcategory = SubCategory.objects.get_or_create(name=str(subcategory_name).title(), category=category)[0] if subcategory_name else None
                        article = ProductArticle.objects.get_or_create(article=str(article_name).upper())[0] if article_name else None

                        # Try to fetch the product, or create if it does not exist
                        product, product_created = Product.objects.get_or_create(
                            barcode=barcode,
                            defaults={
                                'name': product_name,
                                'brand': brand,
                                'section': section,
                                'category': category,
                                'subcategory': subcategory,
                                'is_multi_pack': True if isMultipack == "Yes" else False,
                                'multi_pack_quantity': multipack_qty,
                                'article_no': article,
                            }
                        )

                        # Log field changes for the product
                        updated_fields = []

                        if product_created:
                            product_added_count += 1
                            product_log = "Product added"
                        else:
                            product_log = "Product found (existing)"
                            
                            if product.name != product_name.title():
                                updated_fields.append(f"Name: '{product.name}' -> '{product_name}'")
                                product.name = product_name
                            if product.brand != brand:
                                updated_fields.append(f"Brand: '{product.brand}' -> '{brand}'")
                                product.brand = brand
                            if product.section != section:
                                updated_fields.append(f"Section: '{product.section}' -> '{section}'")
                                product.section = section
                            if product.category != category:
                                updated_fields.append(f"Category: '{product.category}' -> '{category}'")
                                product.category = category
                            if product.subcategory != subcategory:
                                updated_fields.append(f"Subcategory: '{product.subcategory}' -> '{subcategory}'")
                                product.subcategory = subcategory
                            if product.is_multi_pack != (isMultipack == "Yes"):
                                updated_fields.append(f"Is Multi Pack: '{product.is_multi_pack}' -> '{isMultipack == 'Yes'}'")
                                product.is_multi_pack = (isMultipack == "Yes")
                            if product.multi_pack_quantity != multipack_qty:
                                updated_fields.append(f"Multi Pack Quantity: '{product.multi_pack_quantity}' -> '{multipack_qty}'")
                                product.multi_pack_quantity = multipack_qty
                            if product.article_no != article:
                                updated_fields.append(f"Article No: '{product.article_no}' -> '{article}'")
                                product.article_no = article

                            product.save()
                            if updated_fields:
                                product_updated_count += 1
                                product_log += f"; Updated fields: {', '.join(updated_fields)}"

                        # Get or create the size and color
                        size = ProductSize.objects.get_or_create(size=str(size_name).upper())[0] if size_name else None
                        color = ProductColor.objects.get_or_create(color=str(color_name).title())[0] if color_name else None

                        # Convert mfd_date from string to date if necessary
                        if isinstance(mfd_date, str):
                            mfd_date = datetime.strptime(mfd_date, '%Y-%m-%d').date()

                        # Check for existing variant and update or create as needed
                        variant, variant_created = ProductVariant.objects.get_or_create(
                            product=product,
                            mfd_date=mfd_date,
                            defaults={
                                'size': size,
                                'color': color,
                                'buying_price': selling_price * 0.80,
                                'selling_price': selling_price,
                                'applicable_gst': 0,
                                'inventory': inventory * product.multi_pack_quantity if product.is_multi_pack else inventory,
                            }
                        )

                        variant_updated_fields = []
                        if variant_created:
                            variant_added_count += 1
                            variant_log = "Variant added"
                        else:
                            # Log field changes for the variant
                            if variant.color != color:
                                variant_updated_fields.append(f"Color: '{variant.color.color}' -> '{color_name}'")
                                variant.color = color
                            if variant.size != size:
                                variant_updated_fields.append(f"Size: '{variant.size.size}' -> '{size_name}'")
                                variant.selling_price = selling_price
                            if variant.selling_price != selling_price:
                                variant_updated_fields.append(f"Selling Price: '{variant.selling_price}' -> '{selling_price}'")
                                variant.selling_price = selling_price
                            if f"{variant.buying_price}" != f"{(selling_price * 0.80):.2f}":
                                variant_updated_fields.append(f"Buying Price: '{variant.buying_price}' -> '{(selling_price * 0.80):.2f}'")
                                variant.buying_price = selling_price * 0.80
                            if variant.inventory != (int(inventory * product.multi_pack_quantity) if product.is_multi_pack else int(inventory)):
                                variant_updated_fields.append(f"Inventory: '{variant.inventory}' -> '{inventory * product.multi_pack_quantity if product.is_multi_pack else inventory}'")
                                variant.inventory = int(inventory * product.multi_pack_quantity) if product.is_multi_pack else int(inventory)

                            variant.save()
                            if variant_updated_fields:
                                product_updated_count += 1
                                variant_log = f"Variant updated; Updated fields: {', '.join(variant_updated_fields)}"
                            else:
                                variant_log = "Variant updated (no field changes)"


                        success_count += 1
                        success_rows.append((*row, product_log, variant_log, ", ".join(updated_fields + variant_updated_fields)))
                        success_sheet.append((*row, product_log, variant_log, ", ".join(updated_fields + variant_updated_fields)))

                    except ValidationError as e:
                        failure_reason = f"Validation error: {e}"
                        transaction.set_rollback(True)
                        failed_rows.append((*row, failure_reason))
                        error_sheet.append((*row, failure_reason))
                        failure_count += 1
                        if product_created:
                            product_added_count -= 1
                    except Exception as e:
                        failure_reason = f"Error: {e}"
                        transaction.set_rollback(True)
                        failed_rows.append((*row, failure_reason))
                        error_sheet.append((*row, failure_reason))
                        failure_count += 1
                        if product_created:
                            product_added_count -= 1

            # Save the log workbook to a file in the media folder
            failed_imports_path = os.path.join(settings.MEDIA_ROOT, output_file_path)
            log_workbook.save(failed_imports_path)
            
            # Return a JSON response with counts and download URL
            return Response({
                'message': 'File processed with some errors.' if failed_rows else 'File processed successfully.',
                'failed_rows': failed_rows,
                'product_added_count': product_added_count,
                'product_updated_count': product_updated_count,
                'variant_added_count': variant_added_count,
                'variant_updated_count': variant_updated_count,
                'failure_count': failure_count,
                'success_count': success_count,
                'download_url': default_storage.url(output_file_path),
                'transaction_id': db_snapshot_name
            }, status=status.HTTP_207_MULTI_STATUS if failed_rows else status.HTTP_200_OK)

        except Exception as e:
            print(e)
            return JsonResponse({'error': str(e)}, status=500)


class InsertProductRow(APIView):
    permission_classes = [permissions.IsShopOwner]

    def post(self, request):
        with transaction.atomic():
            try:
                barcode = str(request.data.get('barcode')).strip()
                product_name = str(request.data.get('product_name')).strip()
                brand_name = str(request.data.get('brand_name')).strip()
                section_name = str(request.data.get('section_name')).strip()
                article_name = str(request.data.get('article_name')).strip()
                category_name = str(request.data.get('category_name')).strip()
                subcategory_name = str(request.data.get('subcategory_name')).strip()
                color_name = str(request.data.get('color_name')).strip()
                size_name = str(request.data.get('size_name')).strip()
                selling_price = request.data.get('selling_price')
                mfd_date = str(request.data.get('mfd_date')).strip()
                isMultipack = str(request.data.get('isMultipack')).strip()
                inventory = request.data.get('inventory')

                # Retrieve or create the brand, section, category, and subcategory
                brand = Brand.objects.get_or_create(name=str(brand_name).title())[0] if brand_name else None
                section = Section.objects.get_or_create(name=str(section_name).title())[0] if section_name else None
                category = Category.objects.get_or_create(name=str(category_name).title(), section=section)[0] if category_name else None
                subcategory = SubCategory.objects.get_or_create(name=str(subcategory_name).title(), category=category)[0] if subcategory_name else None
                article = ProductArticle.objects.get_or_create(article=str(article_name).upper())[0] if article_name else None
                msg = ""

                # Try to fetch the product, or create if it does not exist
                product, product_created = Product.objects.get_or_create(
                    barcode=barcode,
                    defaults= {
                        'name' : product_name,
                        'brand': brand,
                        'section': section,
                        'category': category,
                        'subcategory': subcategory,
                        'is_multi_pack': True if isMultipack == "Yes" else False,
                        'multi_pack_quantity':  request.data['multipack_qty'],
                        'article_no': article,
                    }
                )

                if product_created:
                    msg += "New Product Created & "
                else:
                    msg += "Fount existing Product & "

                # Get or create the size and color
                size = ProductSize.objects.get_or_create(size=str(size_name).upper())[0] if size_name else None
                color = ProductColor.objects.get_or_create(color=str(color_name).title())[0] if color_name else None
 
                print(selling_price)

                # Check for existing variant and update or create as needed
                variant, variant_created = ProductVariant.objects.update_or_create(
                    product=product,
                    size=size,
                    color=color,
                    mfd_date=mfd_date,
                    defaults={
                        'buying_price': selling_price * 0.80,
                        'selling_price': selling_price,
                        'applicable_gst': 0,
                        'inventory': inventory * product.multi_pack_quantity if product.is_multi_pack else inventory,
                    }
                )

                if variant_created:
                    msg += "New Variant added"
                else:
                    msg += "Updated Variant"
            
                return Response({
                        'message': msg
                    }, status=status.HTTP_200_OK)

            except ValidationError as e:
                failure_reason = f"Validation error: {e}"
                transaction.set_rollback(True)
                return Response({
                        'message': failure_reason
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            except Exception as e:
                failure_reason = f"Error: {e}"
                transaction.set_rollback(True)
                return Response({
                        'message': failure_reason
                    }, status=status.HTTP_400_BAD_REQUEST)
            

class RestoreDBView(APIView):
    permission_classes = [permissions.IsShopOwner]
    """
    Class-based view for restoring a PostgreSQL database from a backup file.
    Only superusers can access this view.
    """

    # @method_decorator(csrf_exempt)
    # @method_decorator(require_http_methods(["POST"]))
    def post(self, request, *args, **kwargs):
        # Get the filename from the request (e.g., passed in JSON body)
        filename = request.data.get('filename')
        if not filename:
            return JsonResponse({
                'status': 'error',
                'message': 'Filename is required.'
            }, status=400)

        # Construct the full path to the backup file
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        backup_file = os.path.join(backup_dir, filename)

        # Check if the backup file exists
        if not os.path.exists(backup_file):
            return JsonResponse({
                'status': 'error',
                'message': f'Backup file {filename} does not exist.'
            }, status=404)

        # Get database configuration from Django settings
        db_name = settings.DATABASES['default']['NAME']
        db_user = settings.DATABASES['default']['USER']
        db_password = settings.DATABASES['default']['PASSWORD']
        db_host = settings.DATABASES['default'].get('HOST', 'localhost')
        db_port = settings.DATABASES['default'].get('PORT', '5432')

        # Prepare the pg_restore command
        pg_restore_command = [
            '/Library/PostgreSQL/16/bin/pg_restore',  # Adjust path if necessary
            '-h', db_host,
            '-p', db_port,
            '-U', db_user,
            '-d', db_name,  # Target database
            '--clean',  # Drop objects before recreating them
            '--if-exists',  # Drop only if objects exist
            backup_file,
        ]

        try:
            # Set environment variable for password
            os.environ['PGPASSWORD'] = db_password
            subprocess.run(pg_restore_command, check=True)
            return JsonResponse({
                'status': 'success',
                'message': f'Database restored from {filename}'
            })
        except subprocess.CalledProcessError as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error restoring database: {str(e)}'
            }, status=500)
        finally:
            # Clean up the environment variable
            del os.environ['PGPASSWORD']



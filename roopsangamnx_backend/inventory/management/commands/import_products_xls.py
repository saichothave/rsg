import openpyxl
from openpyxl import Workbook
from django.core.management.base import BaseCommand, CommandError
from inventory.models import (
    Product, ProductArticle, ProductVariant, ProductColor, ProductSize, Brand, Section, Category, SubCategory
)
from datetime import datetime
from django.core.exceptions import ValidationError
from django.db import transaction
import os
import subprocess
from os import getenv


class Command(BaseCommand):
    help = 'Import product variants from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        failed_rows = []
        output_file_path = "failed_imports.xlsx"

        # Take a snapshot before importing
        try:
            self.stdout.write('Creating database snapshot...')
            subprocess.run([
                'pg_dump', '-U', getenv('DBUSER'), '-h', getenv('DBHOST'), 
                '-Fc', getenv('DBNAME'), '-f', 'snapshot_before_import.dump'
            ], check=True)
            self.stdout.write(self.style.SUCCESS('Database snapshot created successfully'))


            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                # Create a new workbook for failed rows
                error_workbook = Workbook()
                error_sheet = error_workbook.active
                error_sheet.title = "Failed Rows"

                # Copy header row from the original sheet and add a "Failure Reason" column
                headers = [cell.value for cell in sheet[1]]
                headers.append('Failure Reason')
                error_sheet.append(headers)

                # Iterate over rows in the Excel file
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping the header row
                    with transaction.atomic():
                        try:
                            print(row)
                            (barcode, product_name, brand_name, section_name, article_name, 
                            category_name, subcategory_name, color_name, size_name, selling_price, mfd_date, isMultipack, multipack_qty,
                            inventory, isVariant, *cols) = row

                            # Retrieve or create the brand, section, category, and subcategory
                            brand = Brand.objects.get_or_create(name=str(brand_name).title())[0] if brand_name else None
                            section = Section.objects.get_or_create(name=str(section_name).title())[0] if section_name else None
                            category = Category.objects.get_or_create(name=str(category_name).title(), section=section)[0] if category_name else None
                            subcategory = SubCategory.objects.get_or_create(name=str(subcategory_name).title(), category=category)[0] if subcategory_name else None
                            article = ProductArticle.objects.get_or_create(article=str(article_name).upper())[0] if article_name else None

                            # Try to fetch the product, or create if it does not exist
                            product, product_created = Product.objects.get_or_create(
                                barcode=barcode,
                                defaults= {
                                    'name' : product_name,
                                    'brand': brand,
                                    'section': section,
                                    'category': category,
                                    'subcategory': subcategory,
                                    'is_multi_pack': True if isMultipack == "Yes" else False,
                                    'multi_pack_quantity':  multipack_qty,
                                    'article_no': article,
                                }
                            )

                            if product_created:
                                self.stdout.write(self.style.SUCCESS(f"Created new product: {product_name}"))
                            else:
                                self.stdout.write(self.style.SUCCESS(f"Found existing product: {product_name}"))

                            # Get or create the size and color
                            size = ProductSize.objects.get_or_create(size=str(size_name).upper())[0] if size_name else None
                            color = ProductColor.objects.get_or_create(color=str(color_name).title())[0] if color_name else None

                            # Convert mfd_date from string to date if necessary
                            if isinstance(mfd_date, str):
                                mfd_date = datetime.strptime(mfd_date, '%Y-%m-%d').date()

                            # Check for existing variant and update or create as needed
                            variant, variant_created = ProductVariant.objects.update_or_create(
                                product=product,
                                size=size,
                                color=color,
                                mfd_date=mfd_date,
                                defaults={
                                    'buying_price': selling_price * 0.80,
                                    'selling_price': selling_price,
                                    'applicable_gst': 0,
                                    'inventory': inventory * product.multi_pack_quantity if product.is_multi_pack else inventory,
                                }
                            )

                            if variant_created:
                                self.stdout.write(self.style.SUCCESS(f"Created new variant for {product_name} - {mfd_date}"))
                            else:
                                self.stdout.write(self.style.SUCCESS(f"Updated existing variant for {product_name} - {mfd_date}"))

                        except ValidationError as e:
                            failure_reason = f"Validation error: {e}"
                            self.stdout.write(self.style.ERROR(failure_reason))
                            transaction.set_rollback(True)  # Manually rollback for this row
                            failed_rows.append((*row, failure_reason))
                        except Exception as e:
                            failure_reason = f"Error: {e}"
                            self.stdout.write(self.style.ERROR(failure_reason))
                            transaction.set_rollback(True)  # Manually rollback for this row
                            failed_rows.append((*row, failure_reason))

                # Write failed rows to the error workbook
                for failed_row in failed_rows:
                    error_sheet.append(failed_row)

                # Save the error workbook
                error_workbook.save(output_file_path)
                self.stdout.write(self.style.SUCCESS(f"Failed rows saved to {output_file_path}"))

                self.stdout.write(self.style.SUCCESS('Product variants import completed successfully'))

            except FileNotFoundError:
                raise CommandError(f"File '{file_path}' does not exist")
            except Exception as e:
                raise CommandError(f"An error occurred: {e}")
            
        except subprocess.CalledProcessError as e:
            self.stdout.write(self.style.ERROR(f"Snapshot creation failed: {e}"))
